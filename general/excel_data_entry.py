import os
import argparse
from typing import List, Optional
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from utils import (
    validate_filename,
    is_valid_cell_value,
    get_next_empty_row,
    get_worksheet_names,
    parse_data_input,
    find_row_by_phone,
    validate_phone_number
)

class ExcelDataEntry:
    def __init__(self, filename: str = "data.xlsx"):
        self.filename = filename
        self.workbook: Optional[Workbook] = None
        self.current_worksheet: Optional[Worksheet] = None
        self.headers = ["Name", "Phone Number", "Auth", "Summary"]
        self.setup_workbook()

    def setup_workbook(self):
        """Initialize workbook and worksheet."""
        self.load_workbook()
        self.select_worksheet("Sheet1")
        if self.current_worksheet and self.current_worksheet.max_row == 1:
            self.add_headers()
        self.save_workbook()

    def load_workbook(self) -> bool:
        """Load the Excel workbook."""
        try:
            if os.path.exists(self.filename):
                self.workbook = openpyxl.load_workbook(self.filename)
            else:
                self.workbook = openpyxl.Workbook()
            return True
        except InvalidFileException:
            print(f"Error: Invalid Excel file format - {self.filename}")
            return False
        except Exception as e:
            print(f"Error loading workbook: {str(e)}")
            return False

    def save_workbook(self) -> bool:
        """Save the Excel workbook."""
        try:
            if self.workbook is None:
                print("Error: No workbook loaded")
                return False
            self.workbook.save(self.filename)
            return True
        except PermissionError:
            print(f"Error: Unable to save file - Permission denied")
            return False
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            return False

    def select_worksheet(self, worksheet_name: str = "Sheet1") -> bool:
        """Select a worksheet to work with."""
        if self.workbook is None:
            print("Error: No workbook loaded")
            return False

        if worksheet_name in self.workbook.sheetnames:
            self.current_worksheet = self.workbook[worksheet_name]
        else:
            self.current_worksheet = self.workbook.create_sheet(worksheet_name)
        return True

    def add_headers(self) -> bool:
        """Add headers to the worksheet."""
        if self.current_worksheet is None:
            print("Error: No worksheet selected")
            return False

        try:
            for col, header in enumerate(self.headers, 1):
                self.current_worksheet.cell(row=1, column=col, value=header)
            return True
        except Exception as e:
            print(f"Error adding headers: {str(e)}")
            return False

    def add_entry(self, name: str, phone: str, auth: str, summary: str) -> bool:
        """Add or update an entry using the provided values."""
        data = [name, phone, auth, summary]
        if not validate_phone_number(phone):
            print("Error: Invalid phone number format")
            return False

        try:
            if self.current_worksheet is None:
                print("Error: No worksheet selected")
                return False

            existing_row = find_row_by_phone(self.current_worksheet, phone)
            row_num = existing_row if existing_row else get_next_empty_row(self.current_worksheet)

            for col, value in enumerate(data, 1):
                if is_valid_cell_value(value):
                    self.current_worksheet.cell(row=row_num, column=col, value=value)
                else:
                    print(f"Warning: Skipping invalid value: {value}")

            self.save_workbook()
            print("Entry added/updated successfully!")
            return True
        except Exception as e:
            print(f"Error adding/updating entry: {str(e)}")
            return False


e=ExcelDataEntry()
e.add_entry("test","1234567890","test","test")
