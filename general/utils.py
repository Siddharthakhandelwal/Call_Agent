from typing import Any, List, Optional
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import re

def validate_filename(filename: str) -> bool:
    """Validate if the filename has a valid Excel extension."""
    return filename.endswith(('.xlsx', '.xlsm'))

def is_valid_cell_value(value: Any) -> bool:
    """Validate if the value is a supported data type."""
    return isinstance(value, (str, int, float, bool))

def format_cell_reference(row: int, col: int) -> str:
    """Convert row and column numbers to Excel reference (e.g., A1)."""
    return f"{get_column_letter(col)}{row}"

def get_next_empty_row(worksheet: Worksheet) -> int:
    """Find the next empty row in the worksheet."""
    return worksheet.max_row + 1

def validate_column_name(column_name: str) -> bool:
    """Validate if the column name is in correct format (letters only)."""
    return bool(re.match(r'^[A-Za-z]+$', column_name))

def parse_data_input(data_input: str) -> List[str]:
    """Parse comma-separated data input into a list."""
    return [item.strip() for item in data_input.split(',')]

def get_worksheet_names(worksheet_list: List[str]) -> str:
    """Format worksheet names for display."""
    return "\n".join(f"- {name}" for name in worksheet_list)

def find_row_by_phone(worksheet: Worksheet, phone_number: str) -> Optional[int]:
    """Find a row by phone number. Returns row number if found, None otherwise."""
    phone_col = 2  # Phone number is in column B
    for row in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
        cell_value = worksheet.cell(row=row, column=phone_col).value
        if cell_value == phone_number:
            return row
    return None

def validate_phone_number(phone: str) -> bool:
    """Validate phone number format."""
    # Basic validation for demonstration - adjust pattern as needed
    return bool(re.match(r'^\+?1?\d{9,15}$', phone))